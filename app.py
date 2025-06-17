import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import requests

# URL to the Key Excel file in your repo
KEY_FILE_URL = "https://raw.githubusercontent.com/Jobcheruyot/QM_Treasury_Reconcilliations/main/key.xlsx"

st.set_page_config(page_title="Mpesa Reconciliation Web App", layout="wide")
st.title("Mpesa Reconciliation Web App")

st.markdown("""
Upload your **Aspire CSV** and **Safaricom CSV** files below to perform reconciliation.  
All four reports will be included in a single Excel workbook, each on its own sheet.
""")

col1, col2 = st.columns(2)
with col1:
    aspire_file = st.file_uploader("Upload Aspire CSV", type=["csv"], key='aspire')
with col2:
    safaricom_file = st.file_uploader("Upload Safaricom CSV", type=["csv"], key='safaricom')

def style_header(ws, header_row=1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{header_row}"

def autofit(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

if all([aspire_file, safaricom_file]):
    # Load the Aspire and Safaricom files
    aspire = pd.read_csv(aspire_file)
    safaricom = pd.read_csv(safaricom_file)
    # Load the key file from the repo
    key_req = requests.get(KEY_FILE_URL)
    key_req.raise_for_status()
    key = pd.read_excel(io.BytesIO(key_req.content))
    original_safaricom = safaricom.copy()
    original_aspire = aspire.copy()
    original_key = key.copy()

    # ========================
    # PROCESSING (as in your original code)
    # ========================

    # 1. Align safaricom columns if not correct
    if safaricom.shape[1] > 1 and safaricom.columns[0] != 'STORE_NAME':
        new_columns = safaricom.columns[1:].tolist() + ['EXTRA']
        safaricom.columns = new_columns
        safaricom = safaricom.drop(columns='EXTRA', errors='ignore')
    safaricom_cols = [
        'STORE_NAME', 'RECEIPT_NUMBER', 'ACCOUNT_TYPE_NAME', 'TRANSACTION_TYPE', 'START_TIMESTAMP',
        'TRANSACTION_PARTY_DETAILS', 'CREDIT_AMOUNT', 'DEBIT_AMOUNT', 'BALANCE', 'LINKED_TRANSACTION_ID'
    ]
    safaricom = safaricom[[col for col in safaricom_cols if col in safaricom.columns]]

    # 2. Key columns
    key.columns = ['Original_STORE_NAME', 'Clean_STORE_NAME']
    store_map = dict(zip(key['Original_STORE_NAME'], key['Clean_STORE_NAME']))
    safaricom['Store_amend'] = safaricom['STORE_NAME'].map(store_map)

    # 3. Code check
    aspire['TRANSACTION_ID'] = aspire['TRANSACTION_ID'].astype(str)
    safaricom['RECEIPT_NUMBER'] = safaricom['RECEIPT_NUMBER'].astype(str)
    valid_transaction_ids = set(aspire['TRANSACTION_ID'])
    safaricom['code_check'] = safaricom['RECEIPT_NUMBER'].apply(lambda x: x if x in valid_transaction_ids else 'XX')
    safaricom['Ref1'] = safaricom['RECEIPT_NUMBER'].str[:3]
    aspire['Ref1'] = aspire['TRANSACTION_ID'].str[:3]

    # Aspire CODE_VALIDATION
    valid_receipts = set(safaricom['RECEIPT_NUMBER'])
    aspire['CODE_VALIDATION'] = aspire['TRANSACTION_ID'].apply(
        lambda x: 'Yes' if x in valid_receipts else 'No'
    )

    # Store name fix
    aspire['STORE_NAME'] = aspire['STORE_NAME'].replace('OUTERING 2', 'Outering 2')

    # Clean up duplicate TRANSACTION_IDs in aspire
    aspire['TRANSACTION_TYPE'] = aspire['TRANSACTION_TYPE'].fillna('').astype(str).str.strip()
    preferred_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    duplicate_ids = aspire[aspire.duplicated('TRANSACTION_ID', keep=False)]
    duplicate_ids['type_priority'] = duplicate_ids['TRANSACTION_TYPE'].apply(
        lambda x: 0 if x in preferred_types else (1 if x else 2)
    )
    cleaned_duplicates = (
        duplicate_ids.sort_values(by='type_priority')
        .drop_duplicates(subset='TRANSACTION_ID', keep='first')
        .drop(columns='type_priority')
    )
    aspire_nondupes = aspire[~aspire['TRANSACTION_ID'].isin(duplicate_ids['TRANSACTION_ID'])]
    aspire = pd.concat([aspire_nondupes, cleaned_duplicates], ignore_index=True)
    aspire = aspire.sort_values(by='TRANSACTION_ID').reset_index(drop=True)
    aspire = aspire.drop_duplicates(subset='TRANSACTION_ID', keep='first')

    # Validation summary
    code_validation_summary = aspire['CODE_VALIDATION'].value_counts().reset_index()
    code_validation_summary.columns = ['Validation_Status', 'Count']
    total_validation = code_validation_summary['Count'].sum()
    code_validation_summary['Percentage'] = round((code_validation_summary['Count'] / total_validation) * 100, 2)

    # Key for Summary_type
    key = original_key.copy()
    key.columns = ['TRANSACTION_TYPE', 'Summary_type']
    exclude = ['PHONE', 'CHANNEL', 'TRANSACTION_TYPE']
    key = key[~key['TRANSACTION_TYPE'].isin(exclude)]
    key = key.dropna(subset=['TRANSACTION_TYPE'])
    aspire = aspire.merge(key, on='TRANSACTION_TYPE', how='left')

    summary_counts = aspire['Summary_type'].value_counts().reset_index()
    summary_counts.columns = ['Summary_type', 'Count']
    total = summary_counts['Count'].sum()
    summary_counts['Percentage'] = round((summary_counts['Count'] / total) * 100, 2)

    # =======================
    # DAILY SUMMARY SHEET LOGIC
    # =======================
    safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')
    store_summary = pd.DataFrame(safaricom['Store_amend'].dropna().unique(), columns=['Store_amend'])
    store_summary = store_summary.sort_values(by='Store_amend').reset_index(drop=True)

    charge_keywords = [
        'Pay merchant Charge',
        'FSI to Merchant Charge by Receiver',
        'Merchant to Merchant Payment Charge to M-PESA'
    ]
    safaricom_charges = safaricom[safaricom['TRANSACTION_PARTY_DETAILS'].isin(charge_keywords)]
    charges_by_store = safaricom_charges.groupby('Store_amend')['DEBIT_AMOUNT'].sum().reset_index()
    charges_by_store.columns = ['Store_amend', 'Charges']
    store_summary = store_summary.loc[:, ~store_summary.columns.str.contains('Charges', case=False)]
    store_summary1 = store_summary.merge(charges_by_store, on='Store_amend', how='left')
    store_summary1['Charges'] = store_summary1['Charges'].fillna(0)
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Previous day utilization
    aspire['TRANSACTION_TYPE'] = aspire['TRANSACTION_TYPE'].fillna('').astype(str).str.strip()
    preferred_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    dupes = aspire[aspire.duplicated('TRANSACTION_ID', keep=False)]
    dupes['priority'] = dupes['TRANSACTION_TYPE'].apply(lambda x: 0 if x in preferred_types else (1 if x else 2))
    cleaned_dupes = (
        dupes.sort_values(by='priority')
        .drop_duplicates(subset='TRANSACTION_ID', keep='first')
        .drop(columns='priority')
    )
    aspire_unique = aspire[~aspire['TRANSACTION_ID'].isin(dupes['TRANSACTION_ID'])]
    aspire = pd.concat([aspire_unique, cleaned_dupes], ignore_index=True)
    most_common_ref1 = safaricom['Ref1'].mode()[0]
    valid_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    filtered_aspire = aspire[
        (aspire['Ref1'] != most_common_ref1) &
        (aspire['TRANSACTION_TYPE'].isin(valid_types))
    ]
    prev_day_data = filtered_aspire.groupby('STORE_NAME')['AMOUNT'].sum().reset_index()
    prev_day_data.columns = ['Store_amend', 'Prev_day']
    store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Prev_day', case=False)]
    store_summary1 = store_summary1.merge(prev_day_data, on='Store_amend', how='left')
    store_summary1['Prev_day'] = store_summary1['Prev_day'].fillna(0)
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Bank_Transfer
    store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Bank_Transfer|Store_match', case=False)]
    safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')
    bank_transfer_data = safaricom[
        safaricom['TRANSACTION_PARTY_DETAILS'].str.contains(
            'Merchant Account to Organization Settlement Account', case=False, na=False
        )
    ]
    bank_transfer_sum = bank_transfer_data.groupby('Store_amend')['DEBIT_AMOUNT'].sum().reset_index()
    bank_transfer_sum.columns = ['Store_amend', 'Bank_Transfer']
    store_summary1 = store_summary1.merge(bank_transfer_sum, on='Store_amend', how='left')
    store_summary1['Bank_Transfer'] = store_summary1['Bank_Transfer'].fillna(0)
    cols = list(store_summary1.columns)
    cols.remove('Bank_Transfer')
    insert_at = cols.index('Store_amend') + 1
    cols.insert(insert_at, 'Bank_Transfer')
    store_summary1 = store_summary1[cols]
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[store_summary1['Store_amend'] != 'TOTAL'][numeric_cols].sum().to_dict()
    for col, value in totals.items():
        store_summary1.at[store_summary1.index[-1], col] = value

    # Utilized
    common_ref1 = safaricom['Ref1'].mode()[0]
    valid_summaries = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    utilized_data = aspire[
        (aspire['Ref1'] == common_ref1) &
        (aspire['Summary_type'].isin(valid_summaries))
    ]
    utilized_summary = utilized_data.groupby('STORE_NAME')['AMOUNT'].sum().reset_index()
    utilized_summary.columns = ['Store_amend', 'Asp_Utilized']
    store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Asp_Utilized', case=False)]
    store_summary1 = store_summary1.merge(utilized_summary, on='Store_amend', how='left')
    store_summary1['Asp_Utilized'] = store_summary1['Asp_Utilized'].fillna(0)
    cols = list(store_summary1.columns)
    cols.remove('Asp_Utilized')
    insert_at = cols.index('Prev_day') + 1
    cols.insert(insert_at, 'Asp_Utilized')
    store_summary1 = store_summary1[cols]
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[store_summary1['Store_amend'] != 'TOTAL'][numeric_cols].sum().to_dict()
    for col, val in totals.items():
        store_summary1.at[store_summary1.index[-1], col] = val

    # Gross payments (saf_paid)
    safaricom['CREDIT_AMOUNT'] = pd.to_numeric(safaricom['CREDIT_AMOUNT'], errors='coerce')
    saf_paid_data = safaricom[
        safaricom['ACCOUNT_TYPE_NAME'].str.strip().eq('Merchant Account')
    ]
    saf_paid_summary = saf_paid_data.groupby('Store_amend')['CREDIT_AMOUNT'].sum().reset_index()
    saf_paid_summary.columns = ['Store_amend', 'saf_paid']
    store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('saf_paid', case=False)]
    store_summary1 = store_summary1.merge(saf_paid_summary, on='Store_amend', how='left')
    store_summary1['saf_paid'] = store_summary1['saf_paid'].fillna(0)
    cols = list(store_summary1.columns)
    cols.remove('saf_paid')
    insert_at = cols.index('Asp_Utilized') + 1
    cols.insert(insert_at, 'saf_paid')
    store_summary1 = store_summary1[cols]
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Unutilized
    store_summary1['unutilized_txn'] = store_summary1['saf_paid'] - store_summary1['Asp_Utilized']
    cols = list(store_summary1.columns)
    cols.remove('unutilized_txn')
    insert_at = cols.index('saf_paid') + 1
    cols.insert(insert_at, 'unutilized_txn')
    store_summary1 = store_summary1[cols]
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Reversals
    safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')
    reversals_data = safaricom[
        safaricom['LINKED_TRANSACTION_ID'].notna()
    ]
    reversals_sum = reversals_data.groupby('Store_amend')['DEBIT_AMOUNT'].sum().reset_index()
    reversals_sum.columns = ['Store_amend', 'Reversals']
    store_summary1 = store_summary1.merge(reversals_sum, on='Store_amend', how='left')
    store_summary1['Reversals'] = store_summary1['Reversals'].fillna(0)
    cols = list(store_summary1.columns)
    cols.remove('Reversals')
    insert_at = cols.index('unutilized_txn') + 1 if 'unutilized_txn' in cols else len(cols)
    cols.insert(insert_at, 'Reversals')
    store_summary1 = store_summary1[cols]
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Asp_Pending
    aspire['AMOUNT'] = pd.to_numeric(aspire['AMOUNT'], errors='coerce')
    most_common_ref1 = aspire['Ref1'].mode()[0]
    excluded_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    asp_pending_data = aspire[
        (aspire['Ref1'] == most_common_ref1) &
        (~aspire['TRANSACTION_TYPE'].isin(excluded_types))
    ]
    asp_pending_sum = asp_pending_data.groupby('STORE_NAME')['AMOUNT'].sum().reset_index()
    asp_pending_sum.columns = ['Store_amend', 'Asp_Pending']
    store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Asp_Pending', case=False)]
    store_summary1 = store_summary1.merge(asp_pending_sum, on='Store_amend', how='left')
    store_summary1['Asp_Pending'] = store_summary1['Asp_Pending'].fillna(0)
    cols = list(store_summary1.columns)
    cols.remove('Asp_Pending')
    insert_at = cols.index('Reversals') + 1 if 'Reversals' in cols else len(cols)
    cols.insert(insert_at, 'Asp_Pending')
    store_summary1 = store_summary1[cols]
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Unsync
    safaricom['CREDIT_AMOUNT'] = pd.to_numeric(safaricom['CREDIT_AMOUNT'], errors='coerce')
    most_common_ref1 = safaricom['Ref1'].mode()[0]
    unsync_data = safaricom[
        (safaricom['ACCOUNT_TYPE_NAME'] == 'Merchant Account') &
        (safaricom['code_check'] == 'XX') &
        (safaricom['Ref1'] == most_common_ref1)
    ]
    unsync_sum = unsync_data.groupby('Store_amend')['CREDIT_AMOUNT'].sum().reset_index()
    unsync_sum.columns = ['Store_amend', 'unsync']
    store_summary1 = store_summary1.merge(unsync_sum, on='Store_amend', how='left')
    store_summary1['unsync'] = store_summary1['unsync'].fillna(0)
    cols = list(store_summary1.columns)
    if 'unsync' in cols:
        cols.remove('unsync')
        insert_at = cols.index('saf_paid') + 1 if 'saf_paid' in cols else len(cols)
        cols.insert(insert_at, 'unsync')
        store_summary1 = store_summary1[cols]
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

    # Reversal Charges
    cols_to_check = ['unutilized_txn', 'unsync', 'Asp_Pending']
    for col in cols_to_check:
        store_summary1[col] = pd.to_numeric(store_summary1[col], errors='coerce').fillna(0)
    store_summary1['Reversal Charges'] = store_summary1['unutilized_txn'] - store_summary1['unsync'] - store_summary1['Asp_Pending']
    cols = list(store_summary1.columns)
    if 'Reversal Charges' in cols and 'Asp_Pending' in cols:
        cols.remove('Reversal Charges')
        insert_at = cols.index('Asp_Pending') + 1
        cols.insert(insert_at, 'Reversal Charges')
        store_summary1 = store_summary1[cols]

    # Variance
    required_cols = ['unutilized_txn', 'unsync', 'Asp_Pending', 'Reversal Charges']
    for col in required_cols:
        store_summary1[col] = pd.to_numeric(store_summary1[col], errors='coerce').fillna(0)
    store_summary1['Variance'] = (
        store_summary1['unutilized_txn']
        - store_summary1['unsync']
        - store_summary1['Asp_Pending']
        - store_summary1['Reversal Charges']
    )
    cols = list(store_summary1.columns)
    if 'Variance' in cols and 'Reversal Charges' in cols:
        cols.remove('Variance')
        insert_at = cols.index('Reversal Charges') + 1
        cols.insert(insert_at, 'Variance')
        store_summary1 = store_summary1[cols]
    store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
    numeric_cols = store_summary1.select_dtypes(include='number').columns
    totals = store_summary1[numeric_cols].sum().to_dict()
    totals_row = {col: '' for col in store_summary1.columns}
    totals_row.update(totals)
    totals_row['Store_amend'] = 'TOTAL'
    store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)
    desired_order = [
        'Store_amend', 'Bank_Transfer', 'Charges', 'Prev_day',
        'Asp_Utilized', 'saf_paid', 'unutilized_txn', 'unsync',
        'Asp_Pending', 'Reversals', 'Reversal Charges', 'Variance'
    ]
    store_summary1 = store_summary1[desired_order]

    # ====================
    # Generate the reports for each sheet
    # ====================

    # --- 1. Reversals ---
    safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')
    reversal_rows = safaricom[
        safaricom['LINKED_TRANSACTION_ID'].notna() &
        (safaricom['DEBIT_AMOUNT'] > 0)
    ]
    daily_reversals = reversal_rows[['Store_amend', 'TRANSACTION_PARTY_DETAILS', 'DEBIT_AMOUNT', 'LINKED_TRANSACTION_ID']].copy()

    # --- 2. Previous day utilized ---
    aspire['TRANSACTION_TYPE'] = aspire['TRANSACTION_TYPE'].fillna('').astype(str).str.strip()
    preferred_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    dupes = aspire[aspire.duplicated('TRANSACTION_ID', keep=False)]
    dupes['priority'] = dupes['TRANSACTION_TYPE'].apply(lambda x: 0 if x in preferred_types else (1 if x else 2))
    cleaned_dupes = (
        dupes.sort_values(by='priority')
        .drop_duplicates(subset='TRANSACTION_ID', keep='first')
        .drop(columns='priority')
    )
    aspire_unique = aspire[~aspire['TRANSACTION_ID'].isin(dupes['TRANSACTION_ID'])]
    aspire = pd.concat([aspire_unique, cleaned_dupes], ignore_index=True)
    most_common_ref1 = safaricom['Ref1'].mode()[0]
    valid_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    filtered_aspire = aspire[
        (aspire['Ref1'] != most_common_ref1) &
        (aspire['TRANSACTION_TYPE'].isin(valid_types))
    ]
    prev_day_utilized = filtered_aspire[['STORE_NAME', 'TRANSACTION_ID', 'AMOUNT', 'SYSTEM_ENTRY_DATE', 'TRANSACTION_TYPE']]

    # --- 3. Unutilized (Cashed out) ---
    aspire['AMOUNT'] = pd.to_numeric(aspire['AMOUNT'], errors='coerce')
    aspire['SYSTEM_ENTRY_DATE'] = pd.to_datetime(aspire['SYSTEM_ENTRY_DATE'], errors='coerce')
    most_common_ref1 = aspire['Ref1'].mode()[0]
    excluded_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
    asp_pending = aspire[
        (aspire['Ref1'] == most_common_ref1) &
        (~aspire['TRANSACTION_TYPE'].isin(excluded_types))
    ].copy()
    asp_pending['VENDOR_TIME'] = asp_pending['SYSTEM_ENTRY_DATE'].dt.strftime('%H:%M')
    asp_pending['VENDOR_DAY'] = asp_pending['SYSTEM_ENTRY_DATE'].dt.strftime('%d/%m/%Y')
    asp_export = asp_pending[['VENDOR_TIME', 'STORE_NAME', 'TRANSACTION_ID', 'VENDOR_DAY', 'AMOUNT']].copy()
    safaricom['CREDIT_AMOUNT'] = pd.to_numeric(safaricom['CREDIT_AMOUNT'], errors='coerce')
    safaricom['START_TIMESTAMP'] = pd.to_datetime(safaricom['START_TIMESTAMP'], errors='coerce')
    most_common_ref1_saf = safaricom['Ref1'].mode()[0]
    unsync = safaricom[
        (safaricom['ACCOUNT_TYPE_NAME'] == 'Merchant Account') &
        (safaricom['code_check'] == 'XX') &
        (safaricom['Ref1'] == most_common_ref1_saf) &
        (safaricom['LINKED_TRANSACTION_ID'].isna())
    ].copy()
    unsync['VENDOR_TIME'] = unsync['START_TIMESTAMP'].dt.strftime('%H:%M')
    unsync['VENDOR_DAY'] = unsync['START_TIMESTAMP'].dt.strftime('%d/%m/%Y')
    unsync_export = unsync[['VENDOR_TIME', 'Store_amend', 'RECEIPT_NUMBER', 'VENDOR_DAY', 'CREDIT_AMOUNT']].copy()
    unsync_export.columns = ['VENDOR_TIME', 'STORE_NAME', 'TRANSACTION_ID', 'VENDOR_DAY', 'AMOUNT']
    unsync_export = unsync_export[unsync_export['AMOUNT'] > 1]
    store_name_map = dict(zip(original_key.iloc[:, 0], original_key.iloc[:, 1]))
    asp_export['STORE_NAME'] = asp_export['STORE_NAME'].map(store_name_map).fillna(asp_export['STORE_NAME'])
    unsync_export['STORE_NAME'] = unsync_export['STORE_NAME'].map(store_name_map).fillna(unsync_export['STORE_NAME'])
    cashed_out = pd.concat([asp_export, unsync_export], ignore_index=True)
    cashed_out = cashed_out.sort_values(by=['STORE_NAME', 'AMOUNT'], ascending=[True, False])

    # --- 4. Daily Summary (store_summary1) ---
    # Already computed as store_summary1 above

    # --- 5. Prepare workbook ---
    output = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reversals"
    for r in daily_reversals.itertuples(index=False):
        ws1.append(list(r))
    ws1.insert_rows(1)
    ws1.insert_rows(1)
    for i, c in enumerate(list(daily_reversals.columns), 1):
        ws1.cell(row=2, column=i, value=c)
    style_header(ws1, header_row=2)
    autofit(ws1)

    ws2 = wb.create_sheet("Prev_Day_Utilized")
    for r in prev_day_utilized.itertuples(index=False):
        ws2.append(list(r))
    ws2.insert_rows(1)
    ws2.insert_rows(1)
    for i, c in enumerate(list(prev_day_utilized.columns), 1):
        ws2.cell(row=2, column=i, value=c)
    style_header(ws2, header_row=2)
    autofit(ws2)

    ws3 = wb.create_sheet("Cashed_Out")
    for r in cashed_out.itertuples(index=False):
        ws3.append(list(r))
    ws3.insert_rows(1)
    ws3.insert_rows(1)
    for i, c in enumerate(list(cashed_out.columns), 1):
        ws3.cell(row=2, column=i, value=c)
    style_header(ws3, header_row=2)
    autofit(ws3)

    ws4 = wb.create_sheet("Daily_Summary")
    for r in store_summary1.itertuples(index=False):
        ws4.append(list(r))
    ws4.insert_rows(1)
    ws4.insert_rows(1)
    for i, c in enumerate(list(store_summary1.columns), 1):
        ws4.cell(row=2, column=i, value=c)
    style_header(ws4, header_row=2)
    autofit(ws4)

    wb.save(output)
    output.seek(0)

    st.success("Excel workbook with all reports is ready.")
    st.download_button(
        label="Download All Reports (Excel workbook)",
        data=output,
        file_name="Mpesa_Reconciliation_Reports.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.info("Sheets: Reversals, Prev_Day_Utilized, Cashed_Out, Daily_Summary")
else:
    st.info("Please upload both Aspire and Safaricom CSV files to proceed.")
